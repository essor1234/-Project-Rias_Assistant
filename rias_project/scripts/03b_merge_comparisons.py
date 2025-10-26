import sys
import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import time
import os
import argparse # Import argparse for command-line arguments

# ----------------------------------------------------------------------
# Utility: Dual output logging (console + file)
# ----------------------------------------------------------------------
class Tee:
    """Redirect stdout/stderr to both terminal and a log file."""
    def __init__(self, *files):
        self.files = files
    def write(self, obj):
        for f in self.files:
            try: # Add basic error handling for logging
                f.write(obj); f.flush()
            except Exception as e:
                print(f"Error writing to log {getattr(f, 'name', 'unknown')}: {e}", file=sys.__stderr__)
    def flush(self):
        for f in self.files:
            try:
                 f.flush()
            except Exception:
                 pass # Ignore flush errors

# ----------------------------------------------------------------------
# Core Merger Class
# ----------------------------------------------------------------------
class ComparisonMerger:
    """Merges individual comparison Excel files into a single master file."""

    def __init__(self, session_root: Path, template_path: Path):
        self.session_root = session_root
        self.template_path = template_path
        self.output_xlsx = self.session_root / "03_comparison_merged.xlsx"
        self.individual_files_pattern_relative = "*/processed/03_compare_papers_output/*_comparison.xlsx"

        # Setup logging
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.log_dir = self.session_root / "logs"
        self.log_dir.mkdir(parents=True, exist_ok=True)
        self.log_file = self.log_dir / f"merge_compare_log_{timestamp}.txt"
        # Defer logger initialization until needed, avoid issues if run multiple times
        self._logger_initialized = False
        print(f"📄 Merge logs *will* be saved to: {self.log_file}") # Indicate intent

    def _ensure_logging(self):
        """Initializes Tee logger if not already done."""
        if self._logger_initialized:
            return
        try:
            log_f = open(self.log_file, "w", encoding="utf-8")
            # Only replace if not already Tee, otherwise add file
            if not isinstance(sys.stdout, Tee):
                print("(Merge Step) Initializing Tee logger.")
                # Store originals only if we are the first to replace
                if not hasattr(sys, '_original_stdout'): sys._original_stdout = sys.stdout
                if not hasattr(sys, '_original_stderr'): sys._original_stderr = sys.stderr
                sys.stdout = Tee(sys.__stdout__, log_f)
                sys.stderr = Tee(sys.__stderr__, log_f)
            else:
                print("(Merge Step) Adding log file to existing Tee.")
                if log_f not in sys.stdout.files: sys.stdout.files = sys.stdout.files + (log_f,)
                if log_f not in sys.stderr.files: sys.stderr.files = sys.stderr.files + (log_f,)
            self._logger_initialized = True
        except Exception as e:
            print(f"Error initializing logging: {e}", file=sys.__stderr__)


    def find_individual_files(self) -> list:
        """Find all individual comparison files within the session using pathlib.rglob."""
        self._ensure_logging() # Ensure logging is active
        print(f"Searching recursively within {self.session_root} for files matching: '{self.individual_files_pattern_relative}'")
        files = list(self.session_root.rglob(self.individual_files_pattern_relative))
        print(f"Found {len(files)} individual comparison files to merge:")
        files.sort()
        for f in files:
            try: print(f"  - {f.relative_to(self.session_root)}")
            except ValueError: print(f"  - {f}")
        return files

    def read_data_from_excel(self, file_path: Path) -> (pd.DataFrame, pd.DataFrame):
        self._ensure_logging() # Ensure logging is active
        print(f"Reading data from: {file_path.name}")
        try:
            overview_df = pd.read_excel(file_path, sheet_name="Overview", header=0)
            results_df = pd.read_excel(file_path, sheet_name="Results", header=0)
            print(f"  -> Read {len(overview_df)} overview rows, {len(results_df)} results rows.")
            return overview_df, results_df
        except ValueError as ve:
             print(f"⚠️ Sheet missing or other read error in {file_path.name}: {ve}")
             return pd.DataFrame(), pd.DataFrame()
        except Exception as e:
            print(f"⚠️ Failed to read data from {file_path.name}: {e}")
            return pd.DataFrame(), pd.DataFrame()

    def write_merged_to_template(self, overview_df: pd.DataFrame, results_df: pd.DataFrame):
        self._ensure_logging() # Ensure logging is active
        print(f"\n🧾 Writing merged data ({len(overview_df)} overview, {len(results_df)} results) to {self.output_xlsx.name}...")
        try:
            # Check if template exists before loading
            if not self.template_path.exists():
                raise FileNotFoundError(f"Template file not found at: {self.template_path}")

            wb = load_workbook(self.template_path)
            if "Overview" in wb.sheetnames:
                ws = wb["Overview"]; headers = [c.value for c in ws[1]]
                ws.delete_rows(2, ws.max_row) # Clear placeholders
                overview_df_filled = overview_df.fillna('')
                for idx, row in overview_df_filled.iterrows():
                    try: ws.append([row.get(h, "") for h in headers])
                    except Exception as row_err: print(f"   - Error appending overview row {idx}: {row_err}")
                print(f"✅ Merged Overview sheet: Appended {len(overview_df)} entries.")
            else: print("⚠️ 'Overview' sheet not found in template. Skipping.")

            if "Results" in wb.sheetnames:
                ws = wb["Results"]; headers = [c.value for c in ws[1]]
                ws.delete_rows(2, ws.max_row) # Clear placeholders
                results_df_filled = results_df.fillna('')
                for idx, row in results_df_filled.iterrows():
                     try: ws.append([row.get(h, "") for h in headers])
                     except Exception as row_err: print(f"   - Error appending results row {idx}: {row_err}")
                print(f"✅ Merged Results sheet: Appended {len(results_df)} entries.")
            else: print("⚠️ 'Results' sheet not found in template. Skipping.")

            wb.save(self.output_xlsx)
            print(f"💾 Merged comparison Excel saved to: {self.output_xlsx}")
        except Exception as e:
            print(f"❌ Failed to write merged Excel file: {e}"); raise

    def run_merge(self):
        """Main execution logic for merging."""
        self._ensure_logging() # Ensure logging is active for the whole process
        print("\n--- Starting Comparison Merge Process ---")
        individual_files = self.find_individual_files()

        if not individual_files:
            print("⚠️ No individual comparison files found. Creating empty merge file.")
            try:
                if not self.template_path.exists(): raise FileNotFoundError("Template not found")
                wb = load_workbook(self.template_path)
                if "Overview" in wb.sheetnames: wb["Overview"].delete_rows(2, wb["Overview"].max_row)
                if "Results" in wb.sheetnames: wb["Results"].delete_rows(2, wb["Results"].max_row)
                wb.save(self.output_xlsx); print(f"💾 Empty merged file saved: {self.output_xlsx}")
            except Exception as e: print(f"❌ Failed to create empty merged file: {e}")
            return

        all_overview_dfs = []
        all_results_dfs = []
        for file_path in individual_files:
            overview_df, results_df = self.read_data_from_excel(file_path)
            if not overview_df.empty: all_overview_dfs.append(overview_df)
            if not results_df.empty: all_results_dfs.append(results_df)

        if not all_overview_dfs and not all_results_dfs:
            print("⚠️ No valid data found in any individual file. Creating empty merge file.")
            try:
                if not self.template_path.exists(): raise FileNotFoundError("Template not found")
                wb = load_workbook(self.template_path);
                if "Overview" in wb.sheetnames: wb["Overview"].delete_rows(2, wb["Overview"].max_row)
                if "Results" in wb.sheetnames: wb["Results"].delete_rows(2, wb["Results"].max_row)
                wb.save(self.output_xlsx); print(f"💾 Empty merged file saved: {self.output_xlsx}")
            except Exception as e: print(f"❌ Failed to create empty merged file: {e}")
            return

        print(f"Concatenating data from {len(individual_files)} files...")
        combined_overview = pd.concat(all_overview_dfs, ignore_index=True) if all_overview_dfs else pd.DataFrame()
        combined_results = pd.concat(all_results_dfs, ignore_index=True) if all_results_dfs else pd.DataFrame()
        print(f"Total overview rows: {len(combined_overview)}, Total results rows: {len(combined_results)}")

        self.write_merged_to_template(combined_overview, combined_results)
        print("\n✅ Comparison merge finished successfully!")

# ------------------------------------------------------------------
# Bridge function for main.py pipeline - NO LOCK
# ------------------------------------------------------------------
def run(pdf_path, session_dir, prev_results=None):
    """
    Bridge for main.py. Called ONCE at the end by main.py.
    pdf_path is ignored.
    session_dir is the path to the main session folder (e.g., results/SESSION_ID).
    prev_results contains results from all previous steps (not used).
    """
    try:
        session_root = Path(session_dir)
        # Ensure the path passed is actually a directory
        if not session_root.is_dir():
             raise NotADirectoryError(f"Provided session path is not a directory: {session_dir}")

        print(f"--- Running Step 03b: Merge Comparisons for Session {session_root.name} ---")

        # 1. Find the template file
        SCRIPT_DIR = Path(__file__).resolve().parent
        PROJECT_ROOT = SCRIPT_DIR.parent
        template_file = PROJECT_ROOT / "templates" / "Paper_Comparison_Template.xlsx"
        if not template_file.exists():
             raise FileNotFoundError(f"Template file not found at {template_file}")

        # 2. Initialize and run the merger
        merger = ComparisonMerger(session_root=session_root, template_path=template_file)
        merger.run_merge() # Saves file to merger.output_xlsx

        # 3. Return result
        return {
            "status": "success",
            "files": [merger.output_xlsx.name],
            "summary": "individual comparisons merged"
        }

    except Exception as e:
        import traceback
        # Attempt to use logger if initialized, otherwise print directly
        logger = sys.stderr if isinstance(sys.stderr, Tee) else sys.__stderr__
        print(f"ERROR in 03b_merge_comparisons: {e}", file=logger)
        traceback.print_exc(file=logger)
        return {"status": "error", "error": str(e)}

# ----------------------------------------------------------------------
# Updated CLI entry point for direct testing
# ----------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Merge individual paper comparison Excel files from a session directory.")
    parser.add_argument("session_directory", type=str, help="Path to the main session directory (e.g., results/SESSION_ID)")
    args = parser.parse_args()

    session_dir_path = Path(args.session_directory).resolve() # Resolve to absolute path

    print(f"Running merge script in standalone mode for session: {session_dir_path}")

    if not session_dir_path.is_dir():
        print(f"Error: Provided path is not a valid directory: {session_dir_path}")
        sys.exit(1)

    # Assume template is in default location relative to project root
    try:
        project_root = Path(__file__).resolve().parent.parent
        template_path = project_root / "templates" / "Paper_Comparison_Template.xlsx"
        if not template_path.exists():
            print(f"Error: Template file not found at expected location: {template_path}")
            # Try alternative location relative to session_dir? Unlikely structure.
            alt_template = session_dir_path.parent.parent / "templates" / "Paper_Comparison_Template.xlsx"
            if alt_template.exists():
                 print(f"Using alternative template path: {alt_template}")
                 template_path = alt_template
            else:
                 print("Cannot find template file.")
                 sys.exit(1)

        # Instantiate the merger directly for standalone run
        merger_instance = ComparisonMerger(session_root=session_dir_path, template_path=template_path)
        merger_instance.run_merge() # Call the main merge logic

    except Exception as e:
        print(f"\nError during standalone merge test run: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1) # Exit with error code

    print("\n--- Standalone merge script finished ---")

# ----------------------------------------------------------------------