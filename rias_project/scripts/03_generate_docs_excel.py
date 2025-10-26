import json
import sys
import datetime
import time
from pathlib import Path
from dotenv import load_dotenv
from openai import OpenAI
import pandas as pd
from openpyxl import load_workbook
# Removed glob and os imports as locking is removed

# ----------------------------------------------------------------------
# Utility: Dual output logging (console + file)
# ----------------------------------------------------------------------
class Tee:
    """Redirect stdout/stderr to both terminal and a log file."""

    def __init__(self, *files):
        self.files = files

    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush()

    def flush(self):
        for f in self.files:
            f.flush()


# ----------------------------------------------------------------------
# Core class
# ----------------------------------------------------------------------
class DocsExcelGenerator:
    """
    Generates structured Excel reports from LLM analysis of a single text file.
    """

    def __init__(
        self,
        prompt_path: Path,
        template_path: Path,
        output_xlsx_path: Path, # Takes the specific output file path now
        model: str = "gpt-4o",
        max_tokens: int = 10000,
        temperature: float = 0.0,
        max_retries: int = 3,
    ):
        load_dotenv()
        self.prompt_path = Path(prompt_path)
        self.template_path = Path(template_path)
        self.output_xlsx = Path(output_xlsx_path) # Specific output file
        self.output_dir = self.output_xlsx.parent # Get dir from file path
        self.model = model
        self.max_tokens = max_tokens
        self.temperature = temperature
        self.max_retries = max_retries

        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Setup logging directory relative to the output file's location
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        try:
            # Try to place logs in the project's root 'logs' folder if accessible
            project_root_marker = self.output_dir.parent.parent.parent.parent # Heuristic path
            self.log_dir = project_root_marker / "logs"
            if not (project_root_marker / "main.py").exists(): # Check if it looks like project root
                 raise FileNotFoundError("Project root not found, using local logs.")
        except Exception:
            # Fallback to logs dir within the step's output dir
            self.log_dir = self.output_dir / "logs"

        self.log_dir.mkdir(parents=True, exist_ok=True)
        self.log_file = self.log_dir / f"compare_log_{self.output_xlsx.stem}_{timestamp}.txt"
        self._init_logging()
        print(f"📄 Logs will be saved to: {self.log_file}")

    # ------------------------------------------------------------------
    def _init_logging(self):
        log_f = open(self.log_file, "w", encoding="utf-8")
        # Ensure we don't capture logs from parallel runs if Tee is already set
        if not isinstance(sys.stdout, Tee):
            sys.stdout = Tee(sys.__stdout__, log_f)
            sys.stderr = Tee(sys.__stderr__, log_f)
        else:
             # If Tee is already active, just add our log file to it if not present
             if log_f not in sys.stdout.files:
                 sys.stdout.files += (log_f,)
             if log_f not in sys.stderr.files:
                 sys.stderr.files += (log_f,)


    # ------------------------------------------------------------------
    # Utility functions (Unchanged from previous versions)
    # ------------------------------------------------------------------
    @staticmethod
    def truncate_text(text: str, limit: int = 25_000) -> str:
        return text if len(text) <= limit else text[:limit] + "\n\n[Text truncated for LLM]"

    def load_prompt(self) -> str:
        return self.prompt_path.read_text(encoding="utf-8")

    @staticmethod
    def clean_raw(raw: str) -> str:
        if raw.startswith("```"):
            parts = raw.split("```", 2)
            raw = parts[1] if len(parts) > 2 else parts[0]
        raw = raw.strip()
        if raw.lower().startswith("json"):
            raw = raw[4:].strip()
        return raw

    # ------------------------------------------------------------------
    # LLM communication (Unchanged from previous versions)
    # ------------------------------------------------------------------
    def call_llm(self, prompt_text: str) -> str:
        """Send prompt to OpenAI model and return response."""
        client = OpenAI()
        messages = [
            {"role": "system", "content": "You are a research paper analyst. Return ONLY valid JSON following the given schema."},
            {"role": "user", "content": prompt_text},
        ]

        for attempt in range(self.max_retries):
            try:
                resp = client.chat.completions.create(
                    model=self.model,
                    messages=messages,
                    max_tokens=self.max_tokens,
                    temperature=self.temperature,
                    response_format={"type": "json_object"},
                )
                return resp.choices[0].message.content.strip()
            except Exception as e:
                print(f"⚠️ Attempt {attempt + 1} failed: {e}")
                if attempt == self.max_retries - 1:
                    raise
                time.sleep(2 ** attempt)
        return ""

    # ------------------------------------------------------------------
    # Process a single text file (Unchanged logic, just used differently)
    # ------------------------------------------------------------------
    def process_single_paper(self, txt_path: Path, paper_id: str, prompt_base: str):
        """Run LLM extraction for one paper and return two DataFrames."""
        print(f"\n📘 Processing {txt_path.name} ({paper_id}) for individual comparison")
        full_text = txt_path.read_text(encoding="utf-8")
        combined_prompt = prompt_base.replace("<<<DOCUMENT_TEXT>>>", self.truncate_text(full_text))
        raw = self.call_llm(combined_prompt)
        cleaned = self.clean_raw(raw)

        try:
            data = json.loads(cleaned)
            print(f"✅ JSON parsed successfully for {txt_path.name}")
            overview_df = pd.DataFrame(data.get("Overview", []))
            results_df = pd.DataFrame(data.get("Results", []))

            # Ensure PaperID uses the actual file stem
            if not overview_df.empty:
                overview_df["PaperID"] = paper_id
            if not results_df.empty:
                results_df["PaperID"] = paper_id

            return overview_df, results_df
        except json.JSONDecodeError as e:
            print(f"❌ JSON parsing failed for {txt_path.name}: {e}")
            return pd.DataFrame(), pd.DataFrame()

    # ------------------------------------------------------------------
    # Write to Excel template (Unchanged logic, clears placeholders)
    # ------------------------------------------------------------------
    def write_to_template(self, overview_df: pd.DataFrame, results_df: pd.DataFrame):
        """Insert dataframes into a *copy* of the Excel template and save."""
        print(f"\n🧾 Writing results for one paper to {self.output_xlsx.name}...")
        # Load a fresh copy of the template each time
        wb = load_workbook(self.template_path)

        if "Overview" in wb.sheetnames:
            ws = wb["Overview"]
            headers = [cell.value for cell in ws[1]]
            ws.delete_rows(2, ws.max_row) # Clear placeholders
            for _, row in overview_df.iterrows():
                ws.append([row.get(h, "") for h in headers])
            print(f"✅ Overview sheet: {len(overview_df)} entries written.")

        if "Results" in wb.sheetnames:
            ws = wb["Results"]
            headers = [cell.value for cell in ws[1]]
            ws.delete_rows(2, ws.max_row) # Clear placeholders
            for _, row in results_df.iterrows():
                ws.append([row.get(h, "") for h in headers])
            print(f"✅ Results sheet: {len(results_df)} entries written.")

        wb.save(self.output_xlsx)
        print(f"💾 Individual comparison Excel saved to: {self.output_xlsx}")

    # ------------------------------------------------------------------
    # Main execution method for a SINGLE file comparison
    # ------------------------------------------------------------------
    def run_single_comparison(self, txt_file: Path):
        """ Processes ONE txt file and saves its comparison Excel."""
        print(f"\n--- Starting comparison for {txt_file.name} ---")

        prompt_base = self.load_prompt()
        paper_id = txt_file.stem # Use file stem (e.g., 'test3')

        overview_df, results_df = self.process_single_paper(txt_file, paper_id, prompt_base)

        if overview_df.empty and results_df.empty:
            print(f"⚠️ No valid data extracted from {txt_file.name}.")
            # Save an empty file based on the template for consistency
            self.write_to_template(pd.DataFrame(), pd.DataFrame())
        else:
            self.write_to_template(overview_df, results_df)

        print(f"\n✅ Comparison for {txt_file.name} finished!")


# ------------------------------------------------------------------
# Bridge function for main.py pipeline - Runs PER PDF now
# ------------------------------------------------------------------
def run(pdf_path, out_dir, prev=None):
    """
    Bridge function for main.py. Runs comparison for ONE PDF.
    Saves output to the specific out_dir for this step.
    """
    try:
        p_pdf = Path(pdf_path)
        p_out_dir = Path(out_dir)
        pdf_stem = p_pdf.stem

        print(f"--- Running Step 03: Compare Papers for {pdf_stem} ---")

        # 1. Find the corresponding input .txt file from step 01
        # Assumes step 01 output is in '<proc_dir>/01_.../*.txt' relative to out_dir
        processed_dir = p_out_dir.parent
        txt_input_dir = processed_dir / "01_extract_text_output"
        txt_file = txt_input_dir / f"{pdf_stem}.txt"

        if not txt_file.exists():
             raise FileNotFoundError(f"Input text file not found for {pdf_stem} at {txt_file}")

        # 2. Define the output path for THIS specific PDF's comparison
        output_excel_path = p_out_dir / f"{pdf_stem}_comparison.xlsx"

        # 3. Get project paths for template and prompt
        SCRIPT_DIR = Path(__file__).resolve().parent
        PROJECT_ROOT = SCRIPT_DIR.parent
        template_file = PROJECT_ROOT / "templates" / "Paper_Comparison_Template.xlsx"
        prompt_file = PROJECT_ROOT / "prompts" / "[Prompt]compare_prompt.txt"

        # Check required files exist
        if not template_file.exists(): raise FileNotFoundError(f"Template not found: {template_file}")
        if not prompt_file.exists(): raise FileNotFoundError(f"Prompt not found: {prompt_file}")

        # 4. Initialize generator with the SPECIFIC output path
        generator = DocsExcelGenerator(
            prompt_path=prompt_file,
            template_path=template_file,
            output_xlsx_path=output_excel_path # Crucial change
        )

        # 5. Run the comparison for only the current text file
        generator.run_single_comparison(txt_file)

        return {
            "status": "success",
            "files": [output_excel_path.name],
            "summary": f"individual comparison created for {pdf_stem}"
        }

    except Exception as e:
        import traceback
        print(f"ERROR in 03_generate_docs_excel for {pdf_path.stem if pdf_path else 'unknown'}: {e}")
        traceback.print_exc()
        return {"status": "error", "error": str(e)}

# ----------------------------------------------------------------------
# Optional: CLI entry point (if needed for direct testing of single file)
# ----------------------------------------------------------------------
if __name__ == "__main__":
    print("This script is designed to be run via the main pipeline.")
    print("For direct testing, you would need to provide specific paths.")
    # Example (modify paths as needed):
    # test_pdf_path = Path("../results/some_session/test3/raw/test3.pdf")
    # test_out_dir = Path("../results/some_session/test3/processed/03_compare_papers_output")
    # run(test_pdf_path, test_out_dir)

